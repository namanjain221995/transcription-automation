import io
import re
import time
import json
import tempfile
import os
from pathlib import Path
from typing import Optional, Dict, List, Tuple, Any

from dotenv import load_dotenv

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from googleapiclient.errors import HttpError
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

from openai import OpenAI

# =========================
# ENV
# =========================
load_dotenv()

# Non-interactive slot selection (Option B)
# Example: SLOT_CHOICE=2
SLOT_CHOICE = (os.getenv("SLOT_CHOICE") or "").strip()

# Shared drives flag from env (optional)
USE_SHARED_DRIVES = (os.getenv("USE_SHARED_DRIVES") or "").strip().lower() in ("1", "true", "yes", "y")

# OpenAI model override (optional)
DEFAULT_OPENAI_MODEL = (os.getenv("OPENAI_MODEL") or "").strip() or "gpt-5.2-mini"

# =========================
# CONFIG
# =========================
SCOPES = ["https://www.googleapis.com/auth/drive"]
CREDENTIALS_FILE = Path("credentials.json")
TOKEN_FILE = Path("token.json")

# SOURCE: read slots + people from here
ROOT_2026_FOLDER_NAME = "2026"

# DESTINATION: must exist anywhere in Drive. Script will NOT create it.
OUTPUT_ROOT_FOLDER_NAME = "Candidate Result"

FOLDER_MIME = "application/vnd.google-apps.folder"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Deliverable folders (NO "1. Format")
FOLDER_NAMES = [
    "3. Introduction Video",
    "4. Mock Interview (First Call)",
    "5. Project Scenarios",
    "6. 30 Questions Related to Their Niche",
    "7. 50 Questions Related to the Resume",
    "8. Tools & Technology Videos",
]

# Skip these "person" folders under slot
SKIP_PERSON_FOLDERS = {"1. Format"}

# Skip these deliverable folders if they appear under person
SKIP_DELIVERABLE_FOLDERS = {"1. Format"}

# LLM output names with or without ".txt"
LLM_OUTPUT_FILE_RE = re.compile(r"^LLM_OUTPUT__.*(\.txt)?$", re.I)

# Score formats supported:
# - Final Overall Score: 8/10
# - Final Overall Score: 8.5 out of 10
# - Final Overall Score: 60%
# - Final Overall Score: FAIL (30%)
# - Final Overall Score: PASS (80%)
FAIL_PASS_PCT_RX = re.compile(
    r"final\s*overall\s*score\s*[:\-]\s*(pass|fail)\s*(?:\(\s*([0-9]+(?:\.[0-9]+)?)\s*%\s*\)|[:\-]?\s*([0-9]+(?:\.[0-9]+)?)\s*%?)",
    re.I,
)

SCORE_REGEXES = [
    re.compile(
        r"final\s*overall\s*score\s*[:\-]\s*([0-9]+(?:\.[0-9]+)?)\s*(/10|\/\s*10|out\s*of\s*10|%)?",
        re.I,
    ),
    re.compile(
        r"final\s*score\s*[:\-]\s*([0-9]+(?:\.[0-9]+)?)\s*(/10|\/\s*10|out\s*of\s*10|%)?",
        re.I,
    ),
]

MAX_CHARS_TO_MODEL = 80_000

# Percent output clamp
PCT_MIN = 0.0
PCT_MAX = 100.0

SLEEP_BETWEEN_UPLOADS_SEC = 0.25

# Exact Drive file name you want (same for every slot)
OUTPUT_XLSX_NAME = "Deliverables Analysis Sheet.xlsx"

# Give editor access to these emails (folder + xlsx)
EDITOR_EMAILS = [
    "rajvi.patel@techsarasolutions.com",
    "sahil.patel@techsarasolutions.com",
    "soham.piprotar@techsarasolutions.com",
]


# =========================
# ENV + OpenAI
# =========================
def init_openai_client() -> OpenAI:
    api_key = (os.getenv("OPENAI_API_KEY") or "").strip()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY not found in .env or environment variables.")
    return OpenAI(api_key=api_key)


# =========================
# Drive auth
# =========================
def get_drive_service():
    creds = None
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)

    # If token exists but scopes changed, refresh will fail. Force new login.
    if creds and set(creds.scopes or []) != set(SCOPES):
        print("[AUTH] token.json scopes mismatch. Deleting token.json and re-authenticating...")
        TOKEN_FILE.unlink(missing_ok=True)
        creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                print(f"[AUTH] Refresh failed ({e}). Re-authenticating...")
                TOKEN_FILE.unlink(missing_ok=True)
                creds = None

        if not creds or not creds.valid:
            if not CREDENTIALS_FILE.exists():
                raise FileNotFoundError("credentials.json not found next to this script.")
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS_FILE), SCOPES)
            # NOTE: In Docker/headless you may want flow.run_console()
            creds = flow.run_local_server(port=0)
            TOKEN_FILE.write_text(creds.to_json(), encoding="utf-8")

    return build("drive", "v3", credentials=creds)


def _list_kwargs():
    if USE_SHARED_DRIVES:
        return {
            "supportsAllDrives": True,
            "includeItemsFromAllDrives": True,
            "corpora": "allDrives",
        }
    return {}


def _get_media_kwargs():
    if USE_SHARED_DRIVES:
        return {"supportsAllDrives": True}
    return {}


# =========================
# Drive helpers
# =========================
def _escape_drive_q_value(s: str) -> str:
    return s.replace("'", "''")


def drive_list_children(service, parent_id: str, mime_type: Optional[str] = None):
    q_parts = [f"'{parent_id}' in parents", "trashed = false"]
    if mime_type:
        q_parts.append(f"mimeType = '{mime_type}'")
    q = " and ".join(q_parts)

    page_token = None
    while True:
        res = (
            service.files()
            .list(
                q=q,
                fields="nextPageToken, files(id,name,mimeType,modifiedTime,size)",
                pageSize=1000,
                pageToken=page_token,
                **_list_kwargs(),
            )
            .execute()
        )
        for f in res.get("files", []):
            yield f
        page_token = res.get("nextPageToken")
        if not page_token:
            break


def drive_find_child(service, parent_id: str, name: str, mime_type: Optional[str] = None) -> Optional[dict]:
    safe = _escape_drive_q_value(name)
    q_parts = [f"'{parent_id}' in parents", "trashed = false", f"name = '{safe}'"]
    if mime_type:
        q_parts.append(f"mimeType = '{mime_type}'")
    q = " and ".join(q_parts)

    res = (
        service.files()
        .list(
            q=q,
            fields="files(id,name,mimeType,parents,modifiedTime)",
            pageSize=50,
            **_list_kwargs(),
        )
        .execute()
    )
    files = res.get("files", []) or []
    if not files:
        return None
    return sorted(files, key=lambda f: f.get("modifiedTime") or "", reverse=True)[0]


def drive_create_folder(service, parent_id: str, name: str) -> str:
    meta = {"name": name, "mimeType": FOLDER_MIME, "parents": [parent_id]}
    created = service.files().create(body=meta, fields="id", **_list_kwargs()).execute()
    return created["id"]


def drive_download_text(service, file_id: str, mime_type: Optional[str]) -> str:
    fh = io.BytesIO()
    try:
        if mime_type and mime_type.startswith("application/vnd.google-apps."):
            request = service.files().export_media(fileId=file_id, mimeType="text/plain", **_get_media_kwargs())
        else:
            request = service.files().get_media(fileId=file_id, **_get_media_kwargs())

        downloader = MediaIoBaseDownload(fh, request, chunksize=1024 * 1024 * 4)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        fh.seek(0)
        raw = fh.read()
    except HttpError as e:
        print(f"     [WARN] Could not download fileId={file_id}: {e}")
        return ""

    try:
        return raw.decode("utf-8")
    except Exception:
        return raw.decode("latin-1", errors="ignore")


def drive_upload_xlsx(service, parent_id: str, filename: str, local_path: Path) -> str:
    """
    Uploads (create or update) filename inside parent_id.
    Returns the Drive fileId.
    """
    existing = None
    for f in drive_list_children(service, parent_id, None):
        if f.get("name") == filename and f.get("mimeType") != FOLDER_MIME:
            existing = f
            break

    media = MediaFileUpload(str(local_path), mimetype=XLSX_MIME, resumable=True)

    if existing:
        service.files().update(fileId=existing["id"], media_body=media, **_list_kwargs()).execute()
        return existing["id"]
    else:
        meta = {"name": filename, "parents": [parent_id]}
        created = service.files().create(body=meta, media_body=media, fields="id", **_list_kwargs()).execute()
        return created["id"]


def drive_search_folder_anywhere(service, folder_name: str) -> List[dict]:
    safe_name = _escape_drive_q_value(folder_name)
    q = f"name = '{safe_name}' and mimeType = '{FOLDER_MIME}' and trashed=false"
    res = (
        service.files()
        .list(
            q=q,
            fields="files(id,name,parents,modifiedTime)",
            pageSize=200,
            **_list_kwargs(),
        )
        .execute()
    )
    return res.get("files", []) or []


def pick_best_named_folder(candidates: List[dict]) -> dict:
    return sorted(candidates, key=lambda c: (c.get("modifiedTime") or ""), reverse=True)[0]


def drive_grant_editor_access(service, file_id: str, emails: List[str]):
    for email in emails:
        perm = {"type": "user", "role": "writer", "emailAddress": email}
        try:
            service.permissions().create(
                fileId=file_id,
                body=perm,
                sendNotificationEmail=False,
                **_list_kwargs(),
            ).execute()
            print(f"  [PERM] Editor added: {email}")
        except HttpError as e:
            msg = (e.content or b"").decode("utf-8", errors="ignore").lower()
            if "alreadyexists" in msg or "already exists" in msg or "duplicate" in msg:
                print(f"  [PERM] Already has access: {email}")
            else:
                print(f"  [PERM] Failed for {email}: {e}")


# =========================
# SLOT SELECTION (supports SLOT_CHOICE env)
# =========================
def list_slot_folders(service, slots_parent_id: str) -> List[dict]:
    return sorted(
        list(drive_list_children(service, slots_parent_id, FOLDER_MIME)),
        key=lambda x: (x.get("name") or "").lower(),
    )


def choose_slot(service, slots_parent_id: str) -> dict:
    slots = list_slot_folders(service, slots_parent_id)
    if not slots:
        raise RuntimeError("No slot folders found under 2026.")

    if SLOT_CHOICE.isdigit():
        idx = int(SLOT_CHOICE)
        if 1 <= idx <= len(slots):
            chosen = slots[idx - 1]
            print(f"[AUTO] Using SLOT_CHOICE={idx}: {chosen['name']}")
            return chosen
        raise RuntimeError(f"SLOT_CHOICE='{SLOT_CHOICE}' out of range (1..{len(slots)}).")

    # interactive fallback
    print("\n" + "=" * 80)
    print("SELECT SLOT TO PROCESS")
    print("=" * 80)

    for i, s in enumerate(slots, start=1):
        print(f"  {i:2}. {s['name']}")
    print("  EXIT - Exit\n")

    while True:
        choice = input("Choose slot number (e.g. 1) or EXIT: ").strip().lower()
        if choice == "exit":
            raise SystemExit(0)
        if choice.isdigit():
            idx = int(choice)
            if 1 <= idx <= len(slots):
                return slots[idx - 1]
        print("Invalid choice. Try again.")


# =========================
# Score extraction -> PERCENT
# =========================
def clamp_pct(x: Optional[float]) -> Optional[float]:
    if x is None:
        return None
    if x < PCT_MIN or x > PCT_MAX:
        return None
    return x


def normalize_to_percent(value: float, unit: Optional[str]) -> Optional[float]:
    u = (unit or "").lower().strip()

    if "%" in u:
        return value

    if "/10" in u or "out of 10" in u:
        return value * 10.0

    # Heuristic if unit missing:
    if value <= 10.0:
        return value * 10.0
    if 10.0 < value <= 100.0:
        return value

    return None


def extract_score_regex_percent(text: str) -> Optional[float]:
    if not text:
        return None

    m = FAIL_PASS_PCT_RX.search(text)
    if m:
        pct_str = m.group(2) or m.group(3)
        try:
            return clamp_pct(round(float(pct_str), 2))
        except Exception:
            return None

    for rx in SCORE_REGEXES:
        m = rx.search(text)
        if m:
            try:
                val = float(m.group(1))
                unit = m.group(2) if m.lastindex and m.lastindex >= 2 else None
                pct = normalize_to_percent(val, unit)
                return clamp_pct(round(pct, 2)) if pct is not None else None
            except Exception:
                return None

    return None


def extract_score_openai_percent(client: OpenAI, text: str) -> Optional[float]:
    if not text:
        return None

    snippet = text[:MAX_CHARS_TO_MODEL]

    schema = {
        "name": "score_extraction_percent",
        "schema": {
            "type": "object",
            "additionalProperties": False,
            "properties": {
                "score_value": {"type": ["number", "null"]},
                "score_unit": {"type": "string", "enum": ["percent", "out_of_10", "unknown"]},
                "evidence": {"type": "string"},
            },
            "required": ["score_value", "score_unit", "evidence"],
        },
        "strict": True,
    }

    system_prompt = (
        "Extract the FINAL OVERALL SCORE from the given text.\n"
        "Examples:\n"
        "- 'Final Overall Score: 8/10'\n"
        "- 'Final Overall Score: 60%'\n"
        "- 'Final Overall Score: FAIL (30%)'\n"
        "Return score_value and score_unit.\n"
        "If no final score exists, score_value must be null.\n"
        "Do not guess."
    )

    try:
        resp = client.responses.create(
            model=DEFAULT_OPENAI_MODEL,
            input=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": snippet},
            ],
            text={"format": {"type": "json_schema", "json_schema": schema}},
            temperature=0,
        )
    except Exception:
        return None

    try:
        data = json.loads(resp.output_text)
    except Exception:
        return None

    val = data.get("score_value", None)
    unit = data.get("score_unit", "unknown")
    if val is None:
        return None

    try:
        val = float(val)
    except Exception:
        return None

    unit_hint = "%" if unit == "percent" else ("/10" if unit == "out_of_10" else None)
    pct = normalize_to_percent(val, unit_hint)
    return clamp_pct(round(pct, 2)) if pct is not None else None


def pick_best_score_from_llm_outputs_percent(
    service,
    openai_client: OpenAI,
    files: List[dict],
) -> Tuple[Optional[float], Optional[str]]:
    llm_files = [
        f for f in files
        if f.get("mimeType") != FOLDER_MIME and LLM_OUTPUT_FILE_RE.match(f.get("name", ""))
    ]
    if not llm_files:
        return None, None

    # newest first
    llm_files.sort(key=lambda x: (x.get("modifiedTime") or ""), reverse=True)

    for f in llm_files:
        text = drive_download_text(service, f["id"], f.get("mimeType"))

        s = extract_score_regex_percent(text)
        if s is not None:
            return s, f.get("name")

        s = extract_score_openai_percent(openai_client, text)
        if s is not None:
            return s, f.get("name")

    return None, None


# =========================
# Excel creation (PERCENT as text "60%")
# =========================
def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)


def build_slot_workbook_percent(rows: List[Dict[str, Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Deliverables Analysis Sheet"

    headers = ["Person Name"] + FOLDER_NAMES + ["Average % (available)"]
    ws.append(headers)

    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        person = r["person"]
        scores_pct = [r.get(folder) for folder in FOLDER_NAMES]
        numeric = [s for s in scores_pct if isinstance(s, (int, float))]
        avg = round(sum(numeric) / len(numeric), 2) if numeric else None

        row_values = (
            [person]
            + [(f"{s}%" if isinstance(s, (int, float)) else None) for s in scores_pct]
            + ([f"{avg}%" if isinstance(avg, (int, float)) else None])
        )
        ws.append(row_values)

    ws.freeze_panes = "A2"
    autosize_columns(ws)
    return wb


# =========================
# MAIN
# =========================
def main():
    openai_client = init_openai_client()
    service = get_drive_service()

    # --- Find SOURCE 2026 folder (contains slots) ---
    candidates_2026 = drive_search_folder_anywhere(service, ROOT_2026_FOLDER_NAME)
    if not candidates_2026:
        raise RuntimeError(f"Could not find folder '{ROOT_2026_FOLDER_NAME}' anywhere in Drive.")
    base_2026 = pick_best_named_folder(candidates_2026)
    slots_parent_id = base_2026["id"]

    # --- Find OUTPUT ROOT folder in Drive (Candidate Result) ---
    candidates_out = drive_search_folder_anywhere(service, OUTPUT_ROOT_FOLDER_NAME)
    if not candidates_out:
        raise RuntimeError(
            f"Could not find output folder '{OUTPUT_ROOT_FOLDER_NAME}' anywhere in Drive. "
            f"Create it and run again."
        )
    output_root = pick_best_named_folder(candidates_out)
    output_root_id = output_root["id"]

    # --- Choose which slot to process (AUTO if SLOT_CHOICE provided) ---
    slot = choose_slot(service, slots_parent_id)
    slot_name = slot["name"]
    slot_id = slot["id"]

    print(f"\n=== SLOT (SOURCE): {slot_name} ===")

    # --- Create/Use per-slot output folder under Candidate Result ---
    slot_out = drive_find_child(service, output_root_id, slot_name, FOLDER_MIME)
    if not slot_out:
        slot_out_id = drive_create_folder(service, output_root_id, slot_name)
        print(f"[OUTPUT] Created folder: {OUTPUT_ROOT_FOLDER_NAME}/{slot_name}")
    else:
        slot_out_id = slot_out["id"]

    # Give editor access on the per-slot output folder
    print("[PERM] Setting folder editors...")
    drive_grant_editor_access(service, slot_out_id, EDITOR_EMAILS)

    # --- Read people folders from SOURCE slot folder 2026/<SlotName>/ ---
    people_all = sorted(
        list(drive_list_children(service, slot_id, FOLDER_MIME)),
        key=lambda x: (x.get("name") or "").lower(),
    )
    people = [p for p in people_all if (p.get("name") or "") not in SKIP_PERSON_FOLDERS]

    slot_rows: List[Dict[str, Any]] = []

    for person in people:
        person_name = person["name"]
        print(f" - Person: {person_name}")
        row: Dict[str, Any] = {"person": person_name}

        person_child_folders = list(drive_list_children(service, person["id"], FOLDER_MIME))
        folder_map = {
            f["name"]: f
            for f in person_child_folders
            if (f.get("name") or "") not in SKIP_DELIVERABLE_FOLDERS
        }

        for folder_name in FOLDER_NAMES:
            folder_node = folder_map.get(folder_name)
            if not folder_node:
                row[folder_name] = None
                print(f"   * {folder_name}: folder missing")
                continue

            files = list(drive_list_children(service, folder_node["id"], None))
            score_pct, used_filename = pick_best_score_from_llm_outputs_percent(service, openai_client, files)
            row[folder_name] = score_pct

            if score_pct is None:
                print(f"   * {folder_name}: no score found in any LLM_OUTPUT__* file")
            else:
                print(f"   * {folder_name}: {score_pct}% (from {used_filename})")

        slot_rows.append(row)

    wb = build_slot_workbook_percent(slot_rows)

    out_name = OUTPUT_XLSX_NAME

    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        local_xlsx = td / out_name
        wb.save(local_xlsx)

        # Upload into: Drive / Candidate Result / <SlotName> /
        uploaded_file_id = drive_upload_xlsx(service, slot_out_id, out_name, local_xlsx)
        print(f"[OK] Uploaded: {OUTPUT_ROOT_FOLDER_NAME}/{slot_name}/{out_name}")

        # Give editor access on the sheet
        print("[PERM] Setting sheet editors...")
        drive_grant_editor_access(service, uploaded_file_id, EDITOR_EMAILS)

    time.sleep(SLEEP_BETWEEN_UPLOADS_SEC)


if __name__ == "__main__":
    main()

import os
import io
import re
import json
import time
import random
import tempfile
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple

from dotenv import load_dotenv
load_dotenv()

from openpyxl import load_workbook

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from googleapiclient.errors import HttpError
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

# =========================================================
# ENV (Option B)
# =========================================================
SLOT_CHOICE = (os.getenv("SLOT_CHOICE") or "").strip()  # e.g. "2"
USE_SHARED_DRIVES = (os.getenv("USE_SHARED_DRIVES") or "").strip().lower() in ("1", "true", "yes", "y")
HEADLESS_AUTH = (os.getenv("HEADLESS_AUTH") or "").strip().lower() in ("1", "true", "yes", "y")

# Thresholds (you can change in .env too)
MIN_DELIVERABLES = float((os.getenv("MIN_DELIVERABLES") or "55").strip() or 55)
MIN_EYE = float((os.getenv("MIN_EYE") or "70").strip() or 70)
TOP_N = int((os.getenv("TOP_N") or "5").strip() or 5)

# =========================================================
# CONFIG (Drive)
# =========================================================
SCOPES = ["https://www.googleapis.com/auth/drive"]
CREDENTIALS_FILE = Path("credentials.json")
TOKEN_FILE = Path("token.json")

ROOT_2026_FOLDER_NAME = "2025"
OUTPUT_ROOT_FOLDER_NAME = "Candidate Result2"
OUTPUT_XLSX_NAME = "Deliverables Analysis Sheet.xlsx"

FOLDER_MIME = "application/vnd.google-apps.folder"

# =========================================================
# RETRIES
# =========================================================
def execute_with_retries(request, *, max_retries: int = 8, base_sleep: float = 1.0):
    for attempt in range(max_retries):
        try:
            return request.execute()
        except HttpError as e:
            status = getattr(e.resp, "status", None)
            if status in (429, 500, 502, 503, 504):
                if attempt == max_retries - 1:
                    raise
                sleep = (base_sleep * (2 ** attempt)) + random.random()
                print(f"[WARN] Drive transient HTTP {status}. Retry in {sleep:.1f}s")
                time.sleep(sleep)
                continue
            raise

# =========================================================
# Shared Drives kwargs (SAFE per method)
# =========================================================
def _kwargs_for_list() -> Dict[str, Any]:
    # ONLY for files().list()
    if USE_SHARED_DRIVES:
        return {
            "supportsAllDrives": True,
            "includeItemsFromAllDrives": True,
            "corpora": "allDrives",
        }
    return {}

def _kwargs_for_get_media() -> Dict[str, Any]:
    # ONLY for get_media / export_media
    if USE_SHARED_DRIVES:
        return {"supportsAllDrives": True}
    return {}

# =========================================================
# Drive Auth
# =========================================================
def get_drive_service():
    creds = None
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)

    # If token exists but scopes changed, refresh will fail. Force new login.
    if creds and set(creds.scopes or []) != set(SCOPES):
        print("[AUTH] token.json scopes mismatch. Deleting token.json and re-authenticating...")
        TOKEN_FILE.unlink(missing_ok=True)
        creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                print(f"[AUTH] Refresh failed ({e}). Re-authenticating...")
                TOKEN_FILE.unlink(missing_ok=True)
                creds = None

        if not creds or not creds.valid:
            if not CREDENTIALS_FILE.exists():
                raise FileNotFoundError("credentials.json not found next to this script.")
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS_FILE), SCOPES)

            # ✅ server/headless-safe
            if HEADLESS_AUTH or not os.environ.get("DISPLAY"):
                creds = flow.run_console()
            else:
                creds = flow.run_local_server(port=0)

            TOKEN_FILE.write_text(creds.to_json(), encoding="utf-8")

    return build("drive", "v3", credentials=creds)

# =========================================================
# Drive Helpers
# =========================================================
def _escape_drive_q_value(s: str) -> str:
    """
    Google Drive v3 query string escaping:
      - escape backslash first
      - escape single quote as \'
    """
    return s.replace("\\", "\\\\").replace("'", "\\'")

def drive_search_folder_anywhere(service, folder_name: str) -> List[dict]:
    safe_name = _escape_drive_q_value(folder_name)
    q = f"name = '{safe_name}' and mimeType = '{FOLDER_MIME}' and trashed=false"

    out: List[dict] = []
    page_token = None
    while True:
        res = execute_with_retries(
            service.files().list(
                q=q,
                fields="nextPageToken, files(id,name,parents,modifiedTime)",
                pageSize=1000,
                pageToken=page_token,
                **_kwargs_for_list(),
            )
        )
        out.extend(res.get("files", []) or [])
        page_token = res.get("nextPageToken")
        if not page_token:
            break
    return out

def pick_best_named_folder(candidates: List[dict]) -> dict:
    return sorted(candidates, key=lambda c: c.get("modifiedTime") or "", reverse=True)[0]

def drive_list_children(service, parent_id: str, mime_type: Optional[str] = None):
    q_parts = [f"'{parent_id}' in parents", "trashed = false"]
    if mime_type:
        q_parts.append(f"mimeType = '{mime_type}'")
    q = " and ".join(q_parts)

    page_token = None
    while True:
        res = execute_with_retries(
            service.files().list(
                q=q,
                fields="nextPageToken, files(id,name,mimeType,modifiedTime,size)",
                pageSize=1000,
                pageToken=page_token,
                **_kwargs_for_list(),
            )
        )
        for f in res.get("files", []) or []:
            yield f
        page_token = res.get("nextPageToken")
        if not page_token:
            break

def drive_find_child(service, parent_id: str, name: str, mime_type: Optional[str] = None) -> Optional[dict]:
    safe_name = _escape_drive_q_value(name)
    q_parts = [f"'{parent_id}' in parents", "trashed = false", f"name = '{safe_name}'"]
    if mime_type:
        q_parts.append(f"mimeType = '{mime_type}'")
    q = " and ".join(q_parts)

    res = execute_with_retries(
        service.files().list(
            q=q,
            fields="files(id,name,mimeType,modifiedTime)",
            pageSize=50,
            **_kwargs_for_list(),
        )
    )
    files = res.get("files", []) or []
    if not files:
        return None
    return sorted(files, key=lambda f: f.get("modifiedTime") or "", reverse=True)[0]

def drive_download_file(service, file_id: str, dest_path: Path):
    request = service.files().get_media(fileId=file_id, **_kwargs_for_get_media())
    with io.FileIO(dest_path, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request, chunksize=1024 * 1024 * 8)
        done = False
        while not done:
            try:
                _, done = downloader.next_chunk()
            except HttpError as e:
                status = getattr(e.resp, "status", None)
                if status in (429, 500, 502, 503, 504):
                    sleep = 1.0 + random.random()
                    print(f"[WARN] Download transient HTTP {status}. Retry in {sleep:.1f}s")
                    time.sleep(sleep)
                    continue
                raise

# =========================================================
# SLOT SELECTION (Option B)
# =========================================================
def list_slot_folders(service, slots_parent_id: str) -> List[dict]:
    return sorted(
        list(drive_list_children(service, slots_parent_id, FOLDER_MIME)),
        key=lambda x: (x.get("name") or "").lower()
    )

def choose_slot(service, slots_parent_id: str) -> dict:
    slots = list_slot_folders(service, slots_parent_id)
    if not slots:
        raise RuntimeError("No slot folders found under 2026.")

    if SLOT_CHOICE.isdigit():
        idx = int(SLOT_CHOICE)
        if 1 <= idx <= len(slots):
            chosen = slots[idx - 1]
            print(f"[AUTO] Using SLOT_CHOICE={idx}: {chosen['name']}")
            return chosen
        raise RuntimeError(f"SLOT_CHOICE='{SLOT_CHOICE}' out of range (1..{len(slots)}).")

    print("\n" + "=" * 80)
    print("SELECT SLOT TO PROCESS")
    print("=" * 80)
    for i, s in enumerate(slots, start=1):
        print(f"  {i:2}. {s['name']}")
    print("  EXIT - Exit\n")

    while True:
        choice = input("Choose slot number (e.g. 1) or EXIT: ").strip().lower()
        if choice == "exit":
            raise SystemExit(0)
        if choice.isdigit():
            idx = int(choice)
            if 1 <= idx <= len(slots):
                return slots[idx - 1]
        print(" Invalid choice. Try again.")

# =========================================================
# Excel Parsing Helpers
# =========================================================
def _norm_header(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "").strip().lower())

def _parse_pct(cell_val: Any) -> Optional[float]:
    if cell_val is None:
        return None
    if isinstance(cell_val, (int, float)):
        v = float(cell_val)
        if 0 <= v <= 1.0:
            return round(v * 100.0, 2)
        if 0 <= v <= 100.0:
            return round(v, 2)
        return None

    s = str(cell_val).strip()
    if not s:
        return None
    s = s.replace("%", "").strip()

    try:
        v = float(s)
    except Exception:
        return None

    if 0 <= v <= 1.0:
        return round(v * 100.0, 2)
    if 0 <= v <= 100.0:
        return round(v, 2)
    return None

def _find_sheet_by_name_like(wb, keywords: List[str]) -> Optional[str]:
    for name in wb.sheetnames:
        n = name.lower()
        if any(k in n for k in keywords):
            return name
    return None

def _extract_table(ws) -> Tuple[List[str], List[List[Any]]]:
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_norm = [_norm_header(h) for h in header]

    rows = []
    for r in range(2, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue
        rows.append(row_vals)

    return header_norm, rows

def read_deliverables_and_eye_scores_from_xlsx(
    xlsx_path: Path
) -> Tuple[Dict[str, Optional[float]], Dict[str, Optional[float]]]:
    """
    Returns:
      deliverables_avg_by_person: {name: avg_pct or None}
      eye_avg_by_person: {name: eye_pct or None}
    """
    wb = load_workbook(xlsx_path, data_only=True)

    # Deliverables sheet
    if "Deliverables Analysis Sheet" in wb.sheetnames:
        deliver_sheet_name = "Deliverables Analysis Sheet"
    else:
        deliver_sheet_name = _find_sheet_by_name_like(wb, ["deliverables", "analysis sheet"]) or wb.sheetnames[0]

    deliver_ws = wb[deliver_sheet_name]
    d_header, d_rows = _extract_table(deliver_ws)

    # Person column
    person_idx = None
    if "person name" in d_header:
        person_idx = d_header.index("person name")
    else:
        person_idx = next((i for i, h in enumerate(d_header) if "person" in h), None)
    if person_idx is None:
        raise RuntimeError(f"Could not find 'Person Name' column in sheet '{deliver_sheet_name}'.")

    # Average column
    avg_idx = None
    candidates = [
        "average % (available)",
        "average %(available)",
        "average",
        "avg",
    ]
    for target in candidates:
        avg_idx = next((i for i, h in enumerate(d_header) if h == target), None)
        if avg_idx is not None:
            break
    if avg_idx is None:
        avg_idx = next((i for i, h in enumerate(d_header) if "average" in h), None)
    if avg_idx is None:
        raise RuntimeError(f"Could not find deliverables average column in sheet '{deliver_sheet_name}'.")

    # Eye may be in deliverables sheet, but usually it's in a separate "eye ..." sheet
    eye_idx_in_deliver = next((i for i, h in enumerate(d_header) if "eye" in h), None)

    deliverables_avg: Dict[str, Optional[float]] = {}
    eye_avg: Dict[str, Optional[float]] = {}

    for row in d_rows:
        name = str(row[person_idx] or "").strip()
        if not name:
            continue
        deliverables_avg[name] = _parse_pct(row[avg_idx])
        if eye_idx_in_deliver is not None:
            eye_avg[name] = _parse_pct(row[eye_idx_in_deliver])

    # If eye not found/populated in deliverables sheet, parse eye sheet
    needs_eye_sheet = True
    if eye_avg and any(v is not None for v in eye_avg.values()):
        needs_eye_sheet = False

    if needs_eye_sheet:
        eye_sheet_name = _find_sheet_by_name_like(wb, ["eye"])
        if eye_sheet_name:
            eye_ws = wb[eye_sheet_name]
            e_header, e_rows = _extract_table(eye_ws)

            e_person_idx = None
            if "person name" in e_header:
                e_person_idx = e_header.index("person name")
            else:
                e_person_idx = next((i for i, h in enumerate(e_header) if "person" in h), None)

            # Prefer "Average % (available)" in eye sheet too
            e_score_idx = None
            for target in ["average % (available)", "average", "avg"]:
                e_score_idx = next((i for i, h in enumerate(e_header) if h == target), None)
                if e_score_idx is not None:
                    break
            if e_score_idx is None:
                e_score_idx = next((i for i, h in enumerate(e_header) if "average" in h or "score" in h), None)

            if e_person_idx is not None and e_score_idx is not None:
                for row in e_rows:
                    name = str(row[e_person_idx] or "").strip()
                    if not name:
                        continue
                    eye_avg[name] = _parse_pct(row[e_score_idx])

    return deliverables_avg, eye_avg

# =========================================================
# Printing (Table)
# =========================================================
def _pct(v: Optional[float]) -> str:
    return "-" if v is None else f"{v:.2f}%"

def print_table(rows: List[Tuple[str, Optional[float], Optional[float]]], *, title: str):
    rank_w = 4
    name_w = max(len("Person Name"), *(len(r[0]) for r in rows)) if rows else len("Person Name")
    del_w = max(len("Deliverables Avg"), *(len(_pct(r[1])) for r in rows)) if rows else len("Deliverables Avg")
    eye_w = max(len("Eye Moments"), *(len(_pct(r[2])) for r in rows)) if rows else len("Eye Moments")

    print("\n" + title)
    print(f"{'Rank':<{rank_w}}  {'Person Name':<{name_w}}  {'Deliverables Avg':<{del_w}}  {'Eye Moments':<{eye_w}}")
    print(f"{'-'*rank_w}  {'-'*name_w}  {'-'*del_w}  {'-'*eye_w}")
    for i, (name, d, e) in enumerate(rows, start=1):
        print(f"{i:<{rank_w}}  {name:<{name_w}}  {_pct(d):<{del_w}}  {_pct(e):<{eye_w}}")

def combined_rank_score(d: Optional[float], e: Optional[float]) -> float:
    vals = [v for v in (d, e) if isinstance(v, (int, float))]
    if not vals:
        return -1e9
    return sum(vals) / len(vals)

# =========================================================
# MAIN
# =========================================================
def main():
    service = get_drive_service()

    # Find SOURCE root (2025)
    candidates = drive_search_folder_anywhere(service, ROOT_2026_FOLDER_NAME)
    if not candidates:
        raise RuntimeError(f"Could not find folder '{ROOT_2026_FOLDER_NAME}' anywhere in Drive.")
    base_2026 = pick_best_named_folder(candidates)

    # Find OUTPUT root (Candidate Result2)
    out_candidates = drive_search_folder_anywhere(service, OUTPUT_ROOT_FOLDER_NAME)
    if not out_candidates:
        raise RuntimeError(
            f"Could not find output folder '{OUTPUT_ROOT_FOLDER_NAME}' anywhere in Drive. Create it and run again."
        )
    output_root = pick_best_named_folder(out_candidates)

    # Choose slot
    slot = choose_slot(service, base_2026["id"])
    slot_name = slot["name"]
    print(f"\n=== SLOT: {slot_name} ===")

    # Candidate Result2/<Slot>
    slot_out_folder = drive_find_child(service, output_root["id"], slot_name, FOLDER_MIME)
    if not slot_out_folder:
        raise RuntimeError(f"Slot folder not found in '{OUTPUT_ROOT_FOLDER_NAME}': {slot_name}")

    # XLSX file inside slot folder
    xlsx_file = drive_find_child(service, slot_out_folder["id"], OUTPUT_XLSX_NAME, None)
    if not xlsx_file:
        raise RuntimeError(f"Excel not found: {OUTPUT_ROOT_FOLDER_NAME}/{slot_name}/{OUTPUT_XLSX_NAME}")

    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        local_xlsx = td / OUTPUT_XLSX_NAME
        drive_download_file(service, xlsx_file["id"], local_xlsx)

        deliverables_avg, eye_avg = read_deliverables_and_eye_scores_from_xlsx(local_xlsx)

    all_people = sorted(set(deliverables_avg.keys()) | set(eye_avg.keys()), key=lambda s: s.lower())

    rows_all: List[Tuple[str, Optional[float], Optional[float]]] = []
    for p in all_people:
        rows_all.append((p, deliverables_avg.get(p), eye_avg.get(p)))

    # Exact requested line output
    print("\nPerson Name --> Deliverables Average | Eye Moments Score")
    for name, d, e in rows_all:
        print(f"{name} --> {_pct(d)} | {_pct(e)}")

    # Filter by thresholds
    rows_filtered = [
        (n, d, e)
        for (n, d, e) in rows_all
        if (d is not None and d >= MIN_DELIVERABLES) and (e is not None and e >= MIN_EYE)
    ]
    rows_filtered.sort(key=lambda r: combined_rank_score(r[1], r[2]), reverse=True)

    print_table(
        rows_filtered,
        title=f"FILTERED (Deliverables >= {MIN_DELIVERABLES:.0f}% AND Eye >= {MIN_EYE:.0f}%)",
    )

    # Top N
    if rows_filtered:
        top_rows = rows_filtered[:TOP_N]
    else:
        rows_ranked = sorted(rows_all, key=lambda r: combined_rank_score(r[1], r[2]), reverse=True)
        top_rows = rows_ranked[:TOP_N]

    print_table(top_rows, title=f"TOP {TOP_N} (Ranked by BOTH score)")

if __name__ == "__main__":
    main()

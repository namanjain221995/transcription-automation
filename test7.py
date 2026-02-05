import os
import io
import json
import re
import tempfile
from pathlib import Path
from typing import Optional, List, Dict, Any

from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from googleapiclient.errors import HttpError
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

# =========================
# ENV
# =========================
load_dotenv()
SLOT_CHOICE = (os.getenv("SLOT_CHOICE") or "").strip()
USE_SHARED_DRIVES = (os.getenv("USE_SHARED_DRIVES") or "").strip().lower() in ("1", "true", "yes", "y")

# =========================
# CONFIG
# =========================
SCOPES = ["https://www.googleapis.com/auth/drive"]
CREDENTIALS_FILE = Path("credentials.json")
TOKEN_FILE = Path("token.json")

ROOT_2026_FOLDER_NAME = "2026"
OUTPUT_ROOT_FOLDER_NAME = "Candidate Result"
OUTPUT_XLSX_NAME = "Deliverables Analysis Sheet.xlsx"

FOLDER_MIME = "application/vnd.google-apps.folder"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

SKIP_PERSON_FOLDERS = {"1. Format"}

DELIVERABLE_FOLDERS = [
    "3. Introduction Video",
    "4. Mock Interview (First Call)",
    "5. Project Scenarios",
    "6. 30 Questions Related to Their Niche",
    "7. 50 Questions Related to the Resume",
    "8. Tools & Technology Videos",
]

# Script-6 output files look like: <stem>__EYE_result.json
EYE_RESULT_RX = re.compile(r".*__EYE_result\.json$", re.IGNORECASE)

NEW_SHEET_NAME = "eye moments analysis"

# =========================
# Drive auth
# =========================
def get_drive_service():
    creds = None
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)

    # If token exists but scopes changed, refresh will fail. Force new login.
    if creds and set(creds.scopes or []) != set(SCOPES):
        print("[AUTH] token.json scopes mismatch. Deleting token.json and re-authenticating...")
        TOKEN_FILE.unlink(missing_ok=True)
        creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            try:
                creds.refresh(Request())
            except Exception as e:
                print(f"[AUTH] Refresh failed ({e}). Re-authenticating...")
                TOKEN_FILE.unlink(missing_ok=True)
                creds = None

        if not creds or not creds.valid:
            if not CREDENTIALS_FILE.exists():
                raise FileNotFoundError("credentials.json not found next to this script.")
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS_FILE), SCOPES)
            # NOTE: In Docker/headless you may want flow.run_console()
            creds = flow.run_local_server(port=0)
            TOKEN_FILE.write_text(creds.to_json(), encoding="utf-8")

    return build("drive", "v3", credentials=creds)


def _list_kwargs():
    if USE_SHARED_DRIVES:
        return {"supportsAllDrives": True, "includeItemsFromAllDrives": True, "corpora": "allDrives"}
    return {}


def _get_media_kwargs():
    if USE_SHARED_DRIVES:
        return {"supportsAllDrives": True}
    return {}


# =========================
# Drive helpers
# =========================
def _escape_drive_q_value(s: str) -> str:
    return s.replace("'", "''")


def drive_search_folder_anywhere(service, folder_name: str) -> List[dict]:
    safe_name = _escape_drive_q_value(folder_name)
    q = f"name = '{safe_name}' and mimeType = '{FOLDER_MIME}' and trashed=false"
    res = service.files().list(
        q=q, fields="files(id,name,parents,modifiedTime)", pageSize=200, **_list_kwargs()
    ).execute()
    return res.get("files", []) or []


def pick_best_named_folder(candidates: List[dict]) -> dict:
    return sorted(candidates, key=lambda c: (c.get("modifiedTime") or ""), reverse=True)[0]


def drive_list_children(service, parent_id: str, mime_type: Optional[str] = None):
    q_parts = [f"'{parent_id}' in parents", "trashed = false"]
    if mime_type:
        q_parts.append(f"mimeType = '{mime_type}'")
    q = " and ".join(q_parts)

    page_token = None
    while True:
        res = service.files().list(
            q=q,
            fields="nextPageToken, files(id,name,mimeType,modifiedTime,size)",
            pageSize=1000,
            pageToken=page_token,
            **_list_kwargs(),
        ).execute()

        for f in res.get("files", []):
            yield f

        page_token = res.get("nextPageToken")
        if not page_token:
            break


def drive_find_child(service, parent_id: str, name: str, mime_type: Optional[str] = None) -> Optional[dict]:
    safe = _escape_drive_q_value(name)
    q_parts = [f"'{parent_id}' in parents", "trashed = false", f"name = '{safe}'"]
    if mime_type:
        q_parts.append(f"mimeType = '{mime_type}'")
    q = " and ".join(q_parts)

    res = service.files().list(
        q=q, fields="files(id,name,mimeType,parents,modifiedTime)", pageSize=50, **_list_kwargs()
    ).execute()
    files = res.get("files", []) or []
    if not files:
        return None
    return sorted(files, key=lambda f: f.get("modifiedTime") or "", reverse=True)[0]


def drive_download_bytes(service, file_id: str) -> bytes:
    request = service.files().get_media(fileId=file_id, **_get_media_kwargs())
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request, chunksize=1024 * 1024 * 4)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    fh.seek(0)
    return fh.read()


def drive_download_to_path(service, file_id: str, dest_path: Path):
    dest_path.write_bytes(drive_download_bytes(service, file_id))


def drive_upload_xlsx(service, parent_id: str, filename: str, local_path: Path) -> str:
    existing = drive_find_child(service, parent_id, filename, None)
    media = MediaFileUpload(str(local_path), mimetype=XLSX_MIME, resumable=True)

    if existing:
        service.files().update(fileId=existing["id"], media_body=media, **_list_kwargs()).execute()
        return existing["id"]
    else:
        meta = {"name": filename, "parents": [parent_id]}
        created = service.files().create(body=meta, media_body=media, fields="id", **_list_kwargs()).execute()
        return created["id"]


# =========================
# SLOT SELECTION
# =========================
def list_slot_folders(service, slots_parent_id: str) -> List[dict]:
    return sorted(list(drive_list_children(service, slots_parent_id, FOLDER_MIME)),
                  key=lambda x: (x.get("name") or "").lower())


def choose_slot(service, slots_parent_id: str) -> dict:
    slots = list_slot_folders(service, slots_parent_id)
    if not slots:
        raise RuntimeError("No slot folders found under 2026.")

    if SLOT_CHOICE.isdigit():
        idx = int(SLOT_CHOICE)
        if 1 <= idx <= len(slots):
            chosen = slots[idx - 1]
            print(f"[AUTO] Using SLOT_CHOICE={idx}: {chosen['name']}")
            return chosen
        raise RuntimeError(f"SLOT_CHOICE='{SLOT_CHOICE}' out of range (1..{len(slots)}).")

    print("\n" + "=" * 80)
    print("SELECT SLOT TO PROCESS")
    print("=" * 80)
    for i, s in enumerate(slots, start=1):
        print(f"  {i:2}. {s['name']}")
    print("  EXIT - Exit\n")

    while True:
        choice = input("Choose slot number (e.g. 1) or EXIT: ").strip().lower()
        if choice == "exit":
            raise SystemExit(0)
        if choice.isdigit():
            idx = int(choice)
            if 1 <= idx <= len(slots):
                return slots[idx - 1]
        print("Invalid choice. Try again.")


# =========================
# Eye score extraction
# =========================
def safe_float(x) -> Optional[float]:
    try:
        v = float(x)
        if 0.0 <= v <= 100.0:
            return v
        return None
    except Exception:
        return None


def extract_eye_score(raw: bytes) -> Optional[float]:
    try:
        data = json.loads(raw.decode("utf-8", errors="ignore"))
    except Exception:
        return None

    if isinstance(data, dict):
        v = safe_float(data.get("score"))
        if v is not None:
            return v
        # fallback nested
        for k in ("result", "data", "output"):
            if isinstance(data.get(k), dict):
                v2 = safe_float(data[k].get("score"))
                if v2 is not None:
                    return v2
    return None


def avg(values: List[float]) -> Optional[float]:
    vals = [float(v) for v in values if isinstance(v, (int, float))]
    if not vals:
        return None
    return round(sum(vals) / len(vals), 2)


def compute_person_folder_eye_scores(service, person_folder_id: str) -> Dict[str, Optional[float]]:
    """
    Returns {deliverable_folder_name: folder_avg_eye_score_percent_or_None}
    where folder_avg = average(scores from all __EYE_result.json in that folder).
    """
    out: Dict[str, Optional[float]] = {}

    child_folders = list(drive_list_children(service, person_folder_id, FOLDER_MIME))
    folder_map = {f.get("name"): f for f in child_folders if f.get("name")}

    for deliverable in DELIVERABLE_FOLDERS:
        node = folder_map.get(deliverable)
        if not node:
            out[deliverable] = None
            continue

        files = list(drive_list_children(service, node["id"], None))
        result_files = [
            f for f in files
            if f.get("mimeType") != FOLDER_MIME and EYE_RESULT_RX.match((f.get("name") or ""))
        ]

        scores: List[float] = []
        for rf in result_files:
            raw = drive_download_bytes(service, rf["id"])
            s = extract_eye_score(raw)
            if s is not None:
                scores.append(s)

        out[deliverable] = avg(scores)

    return out


# =========================
# Excel helpers
# =========================
def style_header_row(ws):
    for c in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)


def recreate_eye_sheet(wb) -> Any:
    """
    Delete sheet if exists, then create a fresh one.
    """
    if NEW_SHEET_NAME in wb.sheetnames:
        ws_old = wb[NEW_SHEET_NAME]
        wb.remove(ws_old)
    ws = wb.create_sheet(NEW_SHEET_NAME)
    return ws


# =========================
# MAIN
# =========================
def main():
    service = get_drive_service()

    # SOURCE 2026
    candidates_2026 = drive_search_folder_anywhere(service, ROOT_2026_FOLDER_NAME)
    if not candidates_2026:
        raise RuntimeError(f"Could not find folder '{ROOT_2026_FOLDER_NAME}' anywhere in Drive.")
    base_2026 = pick_best_named_folder(candidates_2026)
    slots_parent_id = base_2026["id"]

    # OUTPUT ROOT
    candidates_out = drive_search_folder_anywhere(service, OUTPUT_ROOT_FOLDER_NAME)
    if not candidates_out:
        raise RuntimeError(f"Could not find output folder '{OUTPUT_ROOT_FOLDER_NAME}' anywhere in Drive. Create it and run again.")
    output_root = pick_best_named_folder(candidates_out)
    output_root_id = output_root["id"]

    # Slot
    slot = choose_slot(service, slots_parent_id)
    slot_name = slot["name"]
    slot_id = slot["id"]
    print(f"\n=== SLOT: {slot_name} ===")

    # Output slot folder
    slot_out = drive_find_child(service, output_root_id, slot_name, FOLDER_MIME)
    if not slot_out:
        raise RuntimeError(f"Output slot folder not found: {OUTPUT_ROOT_FOLDER_NAME}/{slot_name} (run test5 first).")
    slot_out_id = slot_out["id"]

    # Find workbook
    sheet_file = drive_find_child(service, slot_out_id, OUTPUT_XLSX_NAME, None)
    if not sheet_file:
        raise RuntimeError(f"Workbook not found: {OUTPUT_ROOT_FOLDER_NAME}/{slot_name}/{OUTPUT_XLSX_NAME}")

    # People (SOURCE)
    people = sorted(
        [
            f for f in drive_list_children(service, slot_id, FOLDER_MIME)
            if (f.get("name") or "").strip() not in SKIP_PERSON_FOLDERS
        ],
        key=lambda x: (x.get("name") or "").lower(),
    )

    # Compute all rows
    rows: List[Dict[str, Any]] = []
    for p in people:
        pname = p["name"]
        print(f" - Person: {pname}")

        folder_scores = compute_person_folder_eye_scores(service, p["id"])
        numeric = [v for v in folder_scores.values() if isinstance(v, (int, float))]
        person_avg = avg(numeric)

        row: Dict[str, Any] = {"Person Name": pname}
        for d in DELIVERABLE_FOLDERS:
            row[d] = folder_scores.get(d)
        row["Average % (available)"] = person_avg
        rows.append(row)

    # Download, update workbook, upload back
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        local_xlsx = td / OUTPUT_XLSX_NAME

        print(f"\n[DL] Downloading workbook: {OUTPUT_XLSX_NAME}")
        drive_download_to_path(service, sheet_file["id"], local_xlsx)

        wb = load_workbook(local_xlsx)
        ws_eye = recreate_eye_sheet(wb)

        headers = ["Person Name"] + DELIVERABLE_FOLDERS + ["Average % (available)"]
        ws_eye.append(headers)
        style_header_row(ws_eye)

        for r in rows:
            vals = [r["Person Name"]]
            for d in DELIVERABLE_FOLDERS:
                v = r.get(d)
                vals.append(f"{round(v, 2)}%" if isinstance(v, (int, float)) else None)
            av = r.get("Average % (available)")
            vals.append(f"{round(av, 2)}%" if isinstance(av, (int, float)) else None)
            ws_eye.append(vals)

        ws_eye.freeze_panes = "A2"
        autosize_columns(ws_eye)

        wb.save(local_xlsx)

        print(f"[UP] Uploading updated workbook with sheet '{NEW_SHEET_NAME}'...")
        drive_upload_xlsx(service, slot_out_id, OUTPUT_XLSX_NAME, local_xlsx)

    print("[DONE] Eye moments analysis sheet created/updated.")


if __name__ == "__main__":
    main()
